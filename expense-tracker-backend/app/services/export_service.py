"""Export service for generating financial reports.

Supports CSV, Excel, and PDF formats with rich formatting.
"""

from datetime import date
from sqlalchemy.orm import Session
from io import BytesIO, StringIO
import csv
from typing import List, Dict

from app.models.expense_models import Expense
from app.models.income_models import Income
from app.schemas.analytics_schema import DashboardOverview
from app.services.analytics_service import AnalyticsService


class ExportService:
    """Service for generating various export formats."""
    
    @staticmethod
    def export_transactions_csv(
        db: Session,
        user_id: int,
        start_date: date,
        end_date: date,
        transaction_type: str = "all"
    ) -> BytesIO:
        """Generate CSV export of transactions.
        
        Args:
            db: Database session
            user_id: User to export
            start_date: Export from date
            end_date: Export to date
            transaction_type: "expense", "income", or "all"
        
        Returns:
            BytesIO object containing CSV data
        """
        output = StringIO()
        writer = csv.writer(output)
        
        # Get transactions
        if transaction_type in ["expense", "all"]:
            expenses = db.query(Expense).filter(
                Expense.user_id == user_id,
                Expense.date >= start_date,
                Expense.date <= end_date
            ).all()
        else:
            expenses = []
        
        if transaction_type in ["income", "all"]:
            incomes = db.query(Income).filter(
                Income.user_id == user_id,
                Income.date >= start_date,
                Income.date <= end_date
            ).all()
        else:
            incomes = []
        
        # Write header
        writer.writerow(["Date", "Type", "Amount", "Category", "Description", "Created"])
        
        # Write expenses
        for expense in expenses:
            writer.writerow([
                expense.date.strftime("%Y-%m-%d"),
                "Expense",
                expense.amount,
                expense.category.name if expense.category else "Uncategorized",
                expense.description or "",
                expense.created_at.strftime("%Y-%m-%d")
            ])
        
        # Write incomes
        for income in incomes:
            writer.writerow([
                income.date.strftime("%Y-%m-%d"),
                "Income",
                income.amount,
                income.category.name if income.category else "Uncategorized",
                income.source or "",
                income.created_at.strftime("%Y-%m-%d")
            ])
        
        # Convert to BytesIO
        output.seek(0)
        bytes_output = BytesIO(output.getvalue().encode('utf-8'))
        bytes_output.seek(0)
        return bytes_output
    
    @staticmethod
    def export_report_csv(
        db: Session,
        user_id: int,
        start_date: date,
        end_date: date
    ) -> BytesIO:
        """Generate financial summary report as CSV.
        
        Args:
            db: Database session
            user_id: User to export
            start_date: Report start
            end_date: Report end
        
        Returns:
            BytesIO object containing report CSV
        """
        output = StringIO()
        writer = csv.writer(output)
        
        # Get overview using the exact export date range
        overview = AnalyticsService.get_dashboard_overview(
            db,
            user_id,
            period="custom",
            start_date=start_date,
            end_date=end_date,
        )
        
        # Write header
        writer.writerow(["Financial Summary Report"])
        writer.writerow([f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"])
        writer.writerow([])
        
        # Write summary
        writer.writerow(["Metric", "Amount"])
        writer.writerow(["Total Income", overview.total_income])
        writer.writerow(["Total Expenses", overview.total_expenses])
        writer.writerow(["Net Savings", overview.net_savings])
        writer.writerow(["Savings Rate (%)", overview.savings_rate])
        writer.writerow(["Average Daily Spending", overview.average_daily_spending])
        writer.writerow(["Average Transaction", overview.average_transaction_amount])
        writer.writerow([])
        
        # Write category breakdown
        writer.writerow(["Category Breakdown"])
        writer.writerow(["Category", "Amount", "% of Total", "Transactions", "Average"])
        for cat in overview.top_expense_categories:
            writer.writerow([
                cat.category_name,
                cat.total_spent,
                cat.percentage_of_total,
                cat.transaction_count,
                cat.average_transaction
            ])
        writer.writerow([])
        
        # Write score
        writer.writerow(["Financial Health Score"])
        writer.writerow(["Overall Score", overview.financial_score.score])
        writer.writerow(["Savings Rate Score", overview.financial_score.savings_rate_score])
        writer.writerow(["Budget Adherence Score", overview.financial_score.budget_adherence_score])
        writer.writerow(["Expense Stability Score", overview.financial_score.expense_stability_score])
        writer.writerow(["Income Stability Score", overview.financial_score.income_stability_score])
        
        output.seek(0)
        bytes_output = BytesIO(output.getvalue().encode('utf-8'))
        bytes_output.seek(0)
        return bytes_output
    
    @staticmethod
    def export_to_excel(
        db: Session,
        user_id: int,
        start_date: date,
        end_date: date
    ) -> BytesIO:
        """Generate Excel export with multiple sheets.
        
        Requires: openpyxl
        
        Sheets:
        - Transactions: All expense/income
        - Summary: Monthly summary
        - Categories: Category breakdown
        - Analytics: Financial metrics
        
        Args:
            db: Database session
            user_id: User to export
            start_date: Export from date
            end_date: Export to date
        
        Returns:
            BytesIO object containing Excel file
        
        Usage:
            excel_file = export_to_excel(db, user_id, date(2026, 1, 1), date(2026, 5, 11))
            with open("report.xlsx", "wb") as f:
                f.write(excel_file.getvalue())
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # TRANSACTIONS SHEET
        ws_trans = wb.create_sheet("Transactions")
        headers = ["Date", "Type", "Amount", "Category", "Description"]
        ws_trans.append(headers)
        
        for header_cell in ws_trans[1]:
            header_cell.font = Font(bold=True)
            header_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_cell.font = Font(bold=True, color="FFFFFF")
        
        expenses = db.query(Expense).filter(
            Expense.user_id == user_id,
            Expense.date >= start_date,
            Expense.date <= end_date
        ).all()
        
        incomes = db.query(Income).filter(
            Income.user_id == user_id,
            Income.date >= start_date,
            Income.date <= end_date
        ).all()
        
        for expense in expenses:
            ws_trans.append([
                expense.date.strftime("%Y-%m-%d"),
                "Expense",
                expense.amount,
                expense.category.name if expense.category else "Uncategorized",
                expense.description or ""
            ])
        
        for income in incomes:
            ws_trans.append([
                income.date.strftime("%Y-%m-%d"),
                "Income",
                income.amount,
                income.category.name if income.category else "Uncategorized",
                income.source or ""
            ])
        
        # SUMMARY SHEET
        ws_summary = wb.create_sheet("Summary")
        overview = AnalyticsService.get_dashboard_overview(
            db,
            user_id,
            period="custom",
            start_date=start_date,
            end_date=end_date,
        )
        
        ws_summary.append(["Metric", "Value"])
        for row in ws_summary[1]:
            row.font = Font(bold=True)
            row.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        ws_summary.append(["Total Income", overview.total_income])
        ws_summary.append(["Total Expenses", overview.total_expenses])
        ws_summary.append(["Net Savings", overview.net_savings])
        ws_summary.append(["Savings Rate %", overview.savings_rate])
        ws_summary.append(["Average Daily Spending", overview.average_daily_spending])
        ws_summary.append(["Average Transaction Amount", overview.average_transaction_amount])
        ws_summary.append(["Total Transactions", overview.total_transactions])
        
        # Convert to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def export_to_pdf(
        db: Session,
        user_id: int,
        start_date: date,
        end_date: date,
        include_charts: bool = False
    ) -> BytesIO:
        """Generate PDF report with financial summary.
        
        Requires: reportlab or weasyprint
        
        Args:
            db: Database session
            user_id: User to export
            start_date: Report start
            end_date: Report end
            include_charts: Include chart images (requires additional setup)
        
        Returns:
            BytesIO object containing PDF file
        
        Usage:
            pdf_file = export_to_pdf(db, user_id, date(2026, 1, 1), date(2026, 5, 11))
            with open("report.pdf", "wb") as f:
                f.write(pdf_file.getvalue())
        
        Note:
            For production use, consider weasyprint for better formatting.
            Current implementation uses reportlab for simpler setup.
        """
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib import colors
        except ImportError:
            raise ImportError("reportlab is required for PDF export. Install with: pip install reportlab")
        
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1f4788'),
            spaceAfter=30,
            alignment=1  # Center
        )
        story.append(Paragraph("Financial Report", title_style))
        story.append(Spacer(1, 0.2 * inch))
        
        # Period
        story.append(Paragraph(
            f"<b>Period:</b> {start_date.strftime('%B %d, %Y')} - {end_date.strftime('%B %d, %Y')}",
            styles['Normal']
        ))
        story.append(Spacer(1, 0.3 * inch))
        
        # Summary
        story.append(Paragraph("<b>Financial Summary</b>", styles['Heading2']))
        overview = AnalyticsService.get_dashboard_overview(
            db,
            user_id,
            period="custom",
            start_date=start_date,
            end_date=end_date,
        )
        
        summary_data = [
            ["Metric", "Amount"],
            ["Total Income", f"${overview.total_income:.2f}"],
            ["Total Expenses", f"${overview.total_expenses:.2f}"],
            ["Net Savings", f"${overview.net_savings:.2f}"],
            ["Savings Rate", f"{overview.savings_rate:.1f}%"],
            ["Average Daily Spending", f"${overview.average_daily_spending:.2f}"],
        ]
        
        summary_table = Table(summary_data, colWidths=[3 * inch, 2 * inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.3 * inch))
        
        # Financial Score
        story.append(Paragraph("<b>Financial Health Score</b>", styles['Heading2']))
        score_data = [
            ["Metric", "Score"],
            ["Overall Score", f"{overview.financial_score.score}/100"],
            ["Savings Rate", f"{overview.financial_score.savings_rate_score}/40"],
            ["Budget Adherence", f"{overview.financial_score.budget_adherence_score}/20"],
            ["Expense Stability", f"{overview.financial_score.expense_stability_score}/30"],
            ["Income Stability", f"{overview.financial_score.income_stability_score}/10"],
        ]
        
        score_table = Table(score_data, colWidths=[3 * inch, 2 * inch])
        score_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(score_table)
        story.append(Spacer(1, 0.3 * inch))
        
        # Top Categories
        story.append(Paragraph("<b>Top Spending Categories</b>", styles['Heading2']))
        cat_data = [["Category", "Amount", "% of Total", "Avg"]]
        for cat in overview.top_expense_categories:
            cat_data.append([
                cat.category_name,
                f"${cat.total_spent:.2f}",
                f"{cat.percentage_of_total:.1f}%",
                f"${cat.average_transaction:.2f}"
            ])
        
        cat_table = Table(cat_data, colWidths=[2 * inch, 1.2 * inch, 1.2 * inch, 1 * inch])
        cat_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(cat_table)
        
        # Build PDF
        doc.build(story)
        output.seek(0)
        return output
